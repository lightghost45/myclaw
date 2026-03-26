import os
import re
import subprocess
from pathlib import Path

import openpyxl

COL_NAME = "本机可用"


def norm(s) -> str:
	if s is None:
		return ""
	s = str(s).strip()
	s = re.sub(r"^[^\w\u4e00-\u9fff]+\s*", "", s) # strip leading emoji/icons
	s = re.sub(r"\s+", " ", s)
	return s.strip()


def parse_ready_set() -> set[str]:
	# openclaw on Windows is typically a PowerShell script (openclaw.ps1),
	# and pwsh may not be installed. Use Windows PowerShell.
	res = subprocess.run(
		["powershell.exe", "-NoProfile", "-Command", "openclaw skills check"],
		capture_output=True,
		text=True,
		encoding="utf-8",
		errors="ignore",
	)
	out = (res.stdout or "") + "\n" + (res.stderr or "")
	if "Ready to use:" not in out or "Missing requirements:" not in out:
		return set()

	ready_block = out.split("Ready to use:",1)[1].split("Missing requirements:",1)[0]
	ready: set[str] = set()
	for line in ready_block.splitlines():
		line = norm(line)
		if line:
			ready.add(line)
	return ready


def main() -> None:
	desktop = Path(os.environ.get("USERPROFILE", "~")).expanduser() / "Desktop"
	xlsx = desktop / "OpenClaw_skills_能力清单.xlsx"
	if not xlsx.exists():
		raise SystemExit(f"File not found: {xlsx}")

	ready = parse_ready_set()

	wb = openpyxl.load_workbook(xlsx)
	ws = wb.active

	header_row =1
	max_col = ws.max_column
	headers = [ws.cell(row=header_row, column=c).value for c in range(1, max_col +1)]

	# add/find target column
	if COL_NAME in headers:
		col_idx = headers.index(COL_NAME) +1
	else:
		col_idx = max_col +1
		ws.cell(row=header_row, column=col_idx, value=COL_NAME)

	# guess skill-name column
	skill_col_candidates = ["Skill", "skill", "技能", "技能名", "名称", "Name", "Skill Name"]
	skill_col = None
	for cand in skill_col_candidates:
		if cand in headers:
			skill_col = headers.index(cand) +1
			break
	if skill_col is None:
		skill_col =1

	# fill1/0
	for r in range(2, ws.max_row +1):
		sval = ws.cell(row=r, column=skill_col).value
		if sval is None or str(sval).strip() == "":
			continue
		n = norm(sval)
		ws.cell(row=r, column=col_idx, value=int(n in ready))

	wb.save(xlsx)
	print(f"Updated {xlsx} (column {col_idx}='{COL_NAME}')")


if __name__ == "__main__":
	main()
